import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Alignment, Border, Side

##################
wb = openpyxl.load_workbook("template.xlsx")
ws = wb["Sheet"]

config_time = "20/11/2001"
config_tong_doanh_thu = "config_tong_doanh_thu"
config_loi_nhuan = "config_loi_nhuan"
config_ti_le_loi_nhuan = "config_ti_le_loi_nhuan"
config_khach_hang_moi = "config_khach_hang_moi"
config_doanh_thu_tb_khach = "config_doanh_thu_tb_khach"

config_doanhThuImg = openpyxl.drawing.image.Image('doanh_thu.png')
config_doanhThuCategoryImg = openpyxl.drawing.image.Image('doanh_thu_cate.png')

bangDoanhThuChiTiet = [
    {
        "ten": "Cà phê",
        "gia": "15000",
        "soLuong": "100",
        "doanhThu": "1500000",
    },
    {
        "ten": "Sữa",
        "gia": "13000",
        "soLuong": "100",
        "doanhThu": "1300000",
    },
    {
        "ten": "Bánh mì",
        "gia": "10000",
        "soLuong": "100",
        "doanhThu": "1000000",
    },
    
]

##################

i = 0
for r in range(1,ws.max_row+1):
    for c in range(1,ws.max_column+1):
        cell = ws.cell(row=r, column=c)
        s = cell.value
        if s != None and "{{time}}" in s: 
            ws.cell(r,c).value = s.replace("{{time}}",config_time) 
            i += 1
        
        if s != None and "{{tong_doanh_thu}}" in s: 
            ws.cell(r,c).value = s.replace("{{tong_doanh_thu}}",config_tong_doanh_thu) 
            i += 1

        if s != None and "{{loi_nhuan}}" in s: 
            ws.cell(r,c).value = s.replace("{{loi_nhuan}}",config_loi_nhuan) 
            i += 1

        if s != None and "{{ti_le_loi_nhuan}}" in s: 
            ws.cell(r,c).value = s.replace("{{ti_le_loi_nhuan}}",config_ti_le_loi_nhuan) 
            i += 1

        if s != None and "{{khách_hang_moi}}" in s: 
            ws.cell(r,c).value = s.replace("{{khách_hang_moi}}",config_khach_hang_moi) 
            i += 1

        if s != None and "{{doanh_thu_tb_khach}}" in s: 
            ws.cell(r,c).value = s.replace("{{doanh_thu_tb_khach}}",config_doanh_thu_tb_khach) 
            i += 1
            

for row_num, item in enumerate(bangDoanhThuChiTiet, start=18):
    ws.cell(row=row_num, column=2, value=row_num - 17)         
    ws.cell(row=row_num, column=3, value=item['ten'])          
    ws.cell(row=row_num, column=4, value=item['gia'])         
    ws.cell(row=row_num, column=5, value=item['soLuong'])     
    ws.cell(row=row_num, column=6, value=item['doanhThu'])    
    
# Apply center alignment to the cells in columns 2 to 6
for row_num in range(18,18 + len(bangDoanhThuChiTiet)):
    for col_num in range(2, 7):
        cell = ws.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create a solid border
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = border


cell_doanhThuImg = ws.cell(row=18 + len(bangDoanhThuChiTiet) + 5, column=2)
config_doanhThuImg.anchor = cell_doanhThuImg.coordinate
ws.add_image(config_doanhThuImg)


cell_doanhThuCategoryImg = ws.cell(row=18 + len(bangDoanhThuChiTiet) + 25, column=2)
config_doanhThuCategoryImg.anchor = cell_doanhThuCategoryImg.coordinate
ws.add_image(config_doanhThuCategoryImg)

wb.save('targetfile.xlsx')
print("{} cells updated".format(i))
